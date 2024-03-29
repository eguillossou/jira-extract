from jira import JIRA
import argparse
import os
import sys
from openpyxl import Workbook
import re
import requests
import json
from datetime import datetime,timedelta

from dateutil.parser import parse

# For windows
# $env:PASSWORD_JIRA='YourPassword'
# python .\jira-extract.py -sp 24
# for debug purpose : import pdb;pdb.set_trace()
# issue_ids_in.fields.__dict__ to have a struct
#pp issue_ids_in.changelog.histories
#pp history.__dict__
def log(str):
    print(str)

#tdc board = 217
TDC_JIRA_BOARD_ID = 217
EXCEL_FILE_NAME = "jira-full-report"
JIRA_URL = "https://jira.talendforge.org/"
TDC_JIRA_SPRINT_PAGINATION = 30

if 'USER_JIRA' not in os.environ:
    log("CONFIG: USER_JIRA environment variable set to default")
USER_LOGIN = os.getenv('USER_JIRA','eguillossou')
PATH_EXCEL_FILE = "c:\\Users\\{}\\".format(USER_LOGIN)

if 'PASSWORD_JIRA' in os.environ:
    USER_PASSWORD = os.environ['PASSWORD_JIRA']
else:
    log("CONFIG: Missing environment variable PASSWORD_JIRA.")
    sys.exit(1)

# New	Accepted
# Accepted	In Progress
# In Progress	Code Review
# Code Review	Validation
# Validation	Merge
# Merge	Final Check
STR_KEY="Issue key"
STR_TYPE="Issue Type"
STR_SUMMARY="Issue Summary"
STR_CREATIONDATE="Creation Date"
STR_RESODATE="Resolution Date"
STR_NEWDATE="In New Date"
STR_CANDIDATEDATE="In Candidate Date"
STR_ACCEPTDATE="In Accepted Date"
STR_PROGRESSDATE="In in progress Date"
STR_REVIEWDATE="In Code review Date"
STR_VALIDDATE="In Validation Date"
STR_MERGEDATE="In Merge Date"
STR_FINALCDATE="In Final check Date"
STR_DONEDATE="In Done Date"
STR_CLOSEDDATE="In Closed Date"
STR_ONHOLDDATE="In On hold Date"
STR_TODODATE="In To Do Date"
STR_BLOCKEDDATE="In Blocked Date"
STR_REJECTEDDATE="In Rejected Date"
STR_NEWTIME="In New Time"
STR_CANDIDATETIME="In Candidate Time"
STR_ACCEPTTIME="In Accepted Time"
STR_PROGRESSTIME="In in progress Time"
STR_REVIEWTIME="In Code review Time"
STR_VALIDTIME="In Validation Time"
STR_MERGETIME="In Merge Time"
STR_FINALCTIME="In Final check Time"
STR_DONETIME="In Done Time"
STR_CLOSEDTIME="In Closed Time"
STR_BLOCKEDTIME="In Blocked Time"
STR_REJECTEDTIME="In Rejected Time"
STR_ONHOLDTIME="In On hold Time"
STR_TODOTIME="In To Do Time"
STR_LEADTIME="Lead time"
STR_BLOCKEDTIME="In Blocked Time"


def arguments():
    parser = argparse.ArgumentParser(description='Launch extraction and process.')
    parser.add_argument('--sp','-sp', type=int,help='Sprint number selection (default active sprint)')
    return parser

def get_active_sprint(handler_jira):
    active_sprint = handler_jira.sprints(TDC_JIRA_BOARD_ID,extended=False, startAt=0, maxResults=1, state="active")
    if(len(active_sprint) ==0):
        active_sprint = handler_jira.sprints(TDC_JIRA_BOARD_ID,extended=False, startAt=0, maxResults=1, state="future")
        print("Future sprint {} selected".format(active_sprint[0].name))
    else:
        print("Active sprint {}".format(active_sprint[0].name))
    return active_sprint[0]

def get_selected_sprint_number(sprint_details_in):
    return int(re.search(r"\d+",str(sprint_details_in)).group())

def get_sprints_list(handler_jira):
    return(handler_jira.sprints(TDC_JIRA_BOARD_ID,extended=True, startAt=TDC_JIRA_SPRINT_PAGINATION))

def get_selected_sprint(sp_nb_in, sprints_list_in):
    return(sprints_list_in[sp_nb_in])

def get_sprint_details(handler_jira, arg_sp , sprints_list_in):
    if(arg_sp is not None):
        # out=[sprints_list_in[i] for i in range(0,len(sprints_list_in)) if sprints_list_in[i].name == 'TDC Sprint {}'.format(arg_sp)]
        for i in range(0,len(sprints_list_in)):
            if sprints_list_in[i].name == 'TDC Sprint {}'.format(arg_sp):
                return(sprints_list_in[i])
        # return(out)
    else:
        return(get_active_sprint(handler_jira))

def get_sprint_start_date(sprint_details_in, field_11070):
    regexp=re.compile('name={}[ ]([0-9]+)[,]startDate=([^,]+)'.format(get_selected_sprint_number(sprint_details_in)))
    
    return re.findall(regexp,str(field_11070))[0].replace("startDate=","")

def construct_jql_query(sp_nb, handler_jira):
    jql_qry='project = TDC AND issuetype in (Bug, "New Feature", "Work Item") AND Sprint = "TDC Sprint {}" ORDER BY labels ASC, RANK'.format(sp_nb)
    return jql_qry

def pad_or_truncate(some_list, target_len):
    return some_list[:target_len] + [0]*(target_len - len(some_list))

def if_already_started(sprints_list_in,sprint_number_in):
    result = re.search(r"\d+",sprints_list_in.split(',')[0])
    if result is not None and int(result.group(0)) < sprint_number_in:
        return(True)
    else:
        return(False)

def if_added_after_started(issue_ids_in, issuelist_added_to_sprint_in):
    return issue_ids_in.key in issuelist_added_to_sprint_in.keys()

def get_reports_from_jira(jira_talend_in, sprint_id_in, user_in, password_in):
    url = '{}rest/greenhopper/1.0/rapid/charts/sprintreport?rapidViewId=217&sprintId={}'.format(jira_talend_in, sprint_id_in)
    headers = {'content-type': 'application/json'}
    r = requests.get(url, headers, auth=(user_in,password_in))
    return(r.json()['contents']['issueKeysAddedDuringSprint'])

def parse_sprints(field_11070):
    return ', '.join(re.findall(r"name=[^,]+",str(field_11070) )).replace("name=","")

def construct_datas(header_list_in, values_issues_in):
    issues_according_to_header_list = [pad_or_truncate(values_issues_in[idx],len(header_list_in)) for idx in range(len(values_issues_in)) ]

    return issues_according_to_header_list

def fillIT(issue_ids_in, sprint_details_in, sprint_number_in, issuelist_added_to_sprint_in):
    customfield_11070 = issue_ids_in.fields.customfield_11070
    sprint_list = parse_sprints(customfield_11070)

    return [issue_ids_in.fields.issuetype.name, 
            issue_ids_in.key,
            issue_ids_in.fields.summary,
            issue_ids_in.fields.customfield_10150,
            issue_ids_in.fields.status.name,
            issue_ids_in.fields.priority.name,
            sprint_list, 
            if_already_started(sprint_list, sprint_number_in),
            if_added_after_started(issue_ids_in, issuelist_added_to_sprint_in),
            ', '.join(issue_ids_in.fields.labels),
            issue_ids_in.fields.customfield_11071]

def fill_cell(ws_in, line_in, col_in, value_in):
    ws_in.cell(row=line_in, column=col_in).value = value_in
    ws.column_dimensions[_get_column_letter(col_in)].width = 25
    # log("fill rowidx=1, colidy={}, value={}".format(col_in, value_in))

def save_file(_allissues, sprint_number_in):
    my_file = "{}{}{}.json".format(PATH_EXCEL_FILE, EXCEL_FILE_NAME, sprint_number_in)
    log("Saving file to : "+my_file)
    with open (my_file, 'a') as f:
        f.write(_allissues)
def save_excel_file(wb_in, sprint_number_in):
    my_file = "{}{}{}.xlsx".format(PATH_EXCEL_FILE, EXCEL_FILE_NAME, sprint_number_in)
    log("Saving file to : "+my_file)
    try:
        wb_in.save(my_file)
    except IOError:
        print("File already opened. Closed it first.")

def _get_column_letter(col_idx):
    """Convert a column number into a column letter (3 -> 'C')

    Right shift the column col_idx by 26 to find column letters in reverse
    order.  These numbers are 1-based, and can be converted to ASCII
    ordinals by adding 64.

    """
    # these indicies correspond to A -> ZZZ and include all allowed
    # columns
    if not 1 <= col_idx <= 18278:
        raise ValueError("Invalid column index {0}".format(col_idx))
    letters = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx, 26)
        # check for exact division and borrow if needed
        if remainder == 0:
            remainder = 26
            col_idx -= 1
        letters.append(chr(remainder+64))
    return ''.join(reversed(letters))
    
# get column number corresponding to header key
def get_col(col_key,dict_in):
    #first column is key of global dict and not values
    col_in=2
    for k,v in dict_in.items():
        if(col_key==k):
            return(col_in)
        col_in+=1
    return(col_in)

def switch_date(transition):
        switcher={
                'To Do':STR_TODODATE,
                'New':STR_NEWDATE,
                'Candidate':STR_CANDIDATEDATE,
                'Accepted':STR_ACCEPTDATE,
                'In Progress':STR_PROGRESSDATE,
                'Code Review':STR_REVIEWDATE,
                'Validation':STR_VALIDDATE,
                'Merge':STR_MERGEDATE,
                'Final Check':STR_FINALCDATE,
                'Done':STR_DONEDATE,
                'Closed':STR_CLOSEDDATE,
                'Blocked':STR_BLOCKEDDATE,
                'Rejected':STR_REJECTEDDATE,
                'On hold':STR_ONHOLDDATE,
            }
        return switcher.get(transition,'\nWARNING: unknown transition date:'+transition+'\n')

def switch_time(transition):
        switcher={
                'To Do':STR_TODOTIME,
                'New':STR_NEWTIME,
                'Candidate':STR_CANDIDATETIME,
                'Accepted':STR_ACCEPTTIME,
                'In Progress':STR_PROGRESSTIME,
                'Code Review':STR_REVIEWTIME,
                'Validation':STR_VALIDTIME,
                'Merge':STR_MERGETIME,
                'Final Check':STR_FINALCTIME,
                'Done':STR_DONETIME,
                'Closed':STR_CLOSEDTIME,
                'Blocked':STR_BLOCKEDTIME,
                'Rejected':STR_REJECTEDTIME,
                'On hold':STR_ONHOLDTIME,
            }
        return switcher.get(transition,'\nWARNING: unknown transition time:'+transition+'\n')

def fill_dataset(issues_in, no_header = False):
    dict_out = {}
    if(not no_header):
        dict_out = {STR_KEY: {} }
        dict_out[STR_KEY]={
                            STR_TYPE:STR_TYPE,
                            STR_SUMMARY:STR_SUMMARY,
                            STR_CREATIONDATE:STR_CREATIONDATE,
                            STR_RESODATE:STR_RESODATE,
                            STR_TODODATE:STR_TODODATE,
                            STR_NEWDATE:STR_NEWDATE,
                            STR_CANDIDATEDATE:STR_CANDIDATEDATE,
                            STR_ACCEPTDATE:STR_ACCEPTDATE,
                            STR_PROGRESSDATE:STR_PROGRESSDATE,
                            STR_REVIEWDATE:STR_REVIEWDATE,
                            STR_VALIDDATE:STR_VALIDDATE,
                            STR_MERGEDATE:STR_MERGEDATE,
                            STR_FINALCDATE:STR_FINALCDATE,
                            STR_DONEDATE:STR_DONEDATE,
                            STR_CLOSEDDATE:STR_CLOSEDDATE,
                            STR_ONHOLDDATE:STR_ONHOLDDATE,
                            STR_BLOCKEDDATE:STR_BLOCKEDDATE,
                            STR_TODOTIME:STR_TODOTIME,
                            STR_NEWTIME:STR_NEWTIME,
                            STR_CANDIDATETIME:STR_CANDIDATETIME,
                            STR_ACCEPTTIME:STR_ACCEPTTIME,
                            STR_PROGRESSTIME:STR_PROGRESSTIME,
                            STR_REVIEWTIME:STR_REVIEWTIME,
                            STR_VALIDTIME:STR_VALIDTIME,
                            STR_MERGETIME:STR_MERGETIME,
                            STR_FINALCTIME:STR_FINALCTIME,
                            STR_DONETIME:STR_DONETIME,
                            STR_CLOSEDTIME:STR_CLOSEDTIME,
                            STR_ONHOLDTIME:STR_ONHOLDTIME,
                            STR_BLOCKEDTIME:STR_BLOCKEDTIME,
                            STR_LEADTIME:STR_LEADTIME
                            }
    for issue in issues_in:
        # import pdb;pdb.set_trace()
        status_update={}
        if hasattr(issue, 'key'):
            key=issue.key
            # print("key: {}".format(key))
            # dict_out = {key:{}}
            dict_out[key] = {}
        else:
            key="empty"
        # Get datetime creation
        datetime_creation = issue.fields.created
        if datetime_creation is not None:
            # Interested in only seconds precision, so slice unnecessary part
            datetime_creation = datetime.strptime(datetime_creation[:19], "%Y-%m-%dT%H:%M:%S")
            dict_out[key][STR_CREATIONDATE] = datetime_creation
            dict_out[key][STR_NEWDATE] = datetime_creation

        # Get datetime resolution
        datetime_resolution = issue.fields.resolutiondate
        if datetime_resolution is not None:
                # Interested in only seconds precision, so slice unnecessary part
            datetime_resolution = datetime.strptime(datetime_resolution[:19], "%Y-%m-%dT%H:%M:%S")
            dict_out[key][STR_RESODATE] = datetime_resolution
            dict_out[key][STR_LEADTIME] = str(datetime_resolution-datetime_creation)
        
        dict_out[key][STR_TYPE] = issue.fields.issuetype.name
        dict_out[key][STR_SUMMARY] = issue.fields.summary

        previous_status_change_date = datetime_creation

        for history in issue.changelog.histories:
            for item in history.items:
                if hasattr(item, 'field') and item.field == "status":
                    # print("status fromString:{}, toString:{}".format(item.fromString,item.toString))
                    if hasattr(history, 'created'):
                        date=datetime.strptime(history.created[:19], "%Y-%m-%dT%H:%M:%S")
                    if item.fromString not in status_update:
                        status_update[item.fromString] = timedelta(0)   
                    if item.toString not in status_update:
                        status_update[item.toString] = timedelta(0)

                    # print(date-previous_status_change_date)

                    dict_out[key][switch_date(item.fromString)] = date
                    # dataset.append([key,issue.fields.issuetype.name if hasattr(issue.fields.issuetype,'name') else "None",item.fromString,item.toString,date])
                    # status_update[item.fromString] += round((date-previous_status_change_date)/3600000 *10) / 10
                    status_update[item.fromString] += date-previous_status_change_date
                    dict_out[key][switch_time(item.fromString)] = str(status_update[item.fromString])
                    previous_status_change_date = date
    return(dict_out)

if __name__ == '__main__':

    if('PASSWORD_JIRA' not in os.environ):
        log("Missing environment variable PASSWORD_JIRA.")
        sys.exit(1)

    jira_options={'server': JIRA_URL ,'agile_rest_path': 'agile'}
    jira=JIRA(options=jira_options,basic_auth=(USER_LOGIN , USER_PASSWORD))

    sprint_list=get_sprints_list(jira)
    sprint_details=get_sprint_details(jira, arguments().parse_args().sp, sprint_list)
    sprint_number=get_selected_sprint_number(sprint_details)
    print("Sprint number : {}".format(sprint_number))

    issues_n=jira.search_issues(construct_jql_query(sprint_number,jira), startAt=0, maxResults=500, validate_query=True, fields=None, expand="changelog", json_result=None)
    datadict = fill_dataset(issues_n)

    if sprint_number>2:
        issues_n_minus_1=jira.search_issues(construct_jql_query(sprint_number-1,jira), startAt=0, maxResults=500, validate_query=True, fields=None, expand="changelog", json_result=None)
        issues_n_minus_2=jira.search_issues(construct_jql_query(sprint_number-2,jira), startAt=0, maxResults=500, validate_query=True, fields=None, expand="changelog", json_result=None)
        datadict_n_minus_1 = fill_dataset(issues_n_minus_1,True)
        datadict_n_minus_2 = fill_dataset(issues_n_minus_2,True)


    if sprint_number>2:
        datadict_merged = {**datadict, **datadict_n_minus_1,**datadict_n_minus_2}
    else:
        datadict_merged = datadict

    wb = Workbook()
    ws = wb.active
    ws.title ="JIRAIssueTransitions"
    ws.freeze_panes = 'B2'
    
    # fill worksheet
    line_in=1
    col_in=1
    max_col=1
    for k,v in datadict_merged.items():
        col_in=1
        fill_cell(ws,line_in,col_in,k)
        col_in +=1
        for sk,sv in v.items():
            fill_cell(ws,line_in,get_col(sk,datadict_merged[STR_KEY]),sv)
            col_in +=1
            if(max_col<col_in): max_col +=1
        line_in +=1

    # ws.auto_filter.ref = 'A1:AF1'
    ws.auto_filter.ref = "A1:{}1".format(_get_column_letter(max_col))

    save_excel_file(wb, sprint_number)
